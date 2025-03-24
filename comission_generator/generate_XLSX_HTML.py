import os
import shutil
from openpyxl import load_workbook

root_folder = "G:\\FIles\\Agents\\February Agents 2025\\"
folder = root_folder + "e"
folder_excels = root_folder + "text_excels"
folder_htmls = root_folder + "text_htmls"


def create_folders(root, dest_excels, dest_htmls):
    try:
        os.makedirs(dest_excels)
        os.makedirs(dest_htmls)
    except:
        print("es")

    for dirpath, dirnames, filenames in os.walk(root):
        for filename in filenames:
            source_file = os.path.join(dirpath, filename)
            destination_file = os.path.join(dest_excels, filename)
            try:
                shutil.copy2(source_file, destination_file)
            except:
                pass
            print(filename)


def rename_files(root_folder):
    """
    Renames all files in subfolders of the given root folder with the format name__key____month___.xlsx
    to key_month.xlsx
    """
    for dirpath, dirnames, filenames in os.walk(root_folder):
        for filename in filenames:
            if filename.endswith('.xlsx') and 'V00' or 'K00' in filename:
                names = filename.split(' ')
                name, key, month = '', '', ''
                # if 'Report' in names:
                for i in names:
                    if 'V00' in i or 'K00' in i:
                        key = i
                    if i == 'Feb':
                        month = 'February'
                    if i == 'Jan':
                        month = 'January'
                    if i == 'Mar':
                        month = 'March'
                    if i == 'Apr':
                        month = 'April'
                    if i == 'May':
                        month = 'May'
                    if i == 'Jun':
                        month = 'June'
                    if i == 'July' or i == "Jul":
                        month = 'July'
                    if i == 'Aug' or i == "Aug.xlsx":
                        month = 'August'
                    if i == 'Sep':
                        month = 'September'
                    if i == 'Oct':
                        month = 'October'
                    if i == 'Nov':
                        month = 'November'
                    if i == 'Dec':
                        month = 'December'
                    if i == 'First':
                        month = 'First'
                    if i == 'Second' or i == '2':
                        month = 'Second'
                    if i == 'Third' or i == '2':
                        month = 'Third'
                    if i == 'Fourth' or i == '2':
                        month = 'Fourth'
                    if i == 'Total' or i == 'Total.xlsx':
                        month = 'Total'

                month = month.replace('_.xlsx', '')
                print(month)
                new_filename = f"{key}_2025_{month}.xlsx"
                try:
                    os.rename(os.path.join(dirpath, filename), os.path.join(dirpath, new_filename))
                except:
                    print(new_filename)
                    pass


def create_tables(name_html, name_destination):
    link = 'ftp://oliverweber%40agentsoliverweber.com:Zh5%5DMVF%28GhZ%7B@server1.agentsoliverweber.com:21//ftp/' + f'{name_html}.xlsx'

    # with urllib.request.urlopen(link) as response, open(f"{name_html}.xlsx", 'wb') as out_file:
    #     data = response.read()  # Read the contents of the file
    #     out_file.write(data)  # Write the contents to the local file

    try:
        wb = load_workbook(f'{name_html}.xlsx')
        ws = ""
        try:
            ws = wb.active
        except:
            ws = wb['Аркуш1']

        html = """{%extends 'base.html'%}
                            {% block content %}     
        <meta charset="utf-8">
        <link rel="stylesheet" href="https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/css/bootstrap.min.css" integrity="sha384-Gn5384xqQ1aoWXA+058RXPxPg6fy4IWvTNh0E263XmFcJlSAwiGgFAW/dAiS6JXm" crossorigin="anonymous">
    <!--<link rel="stylesheet" type="text/css" href="normas.css">-->
    <style>
        body{
            background-color: #d6e1f8;

        }
        th, td {
            border: 1px solid #ccc;

            text-align: center;
        }

        table {

            border-collapse: collapse !important; /* optional: collapse the table borders */
        }

        table tr {
            background-color: #8BC34A !important; /* Set background color of all rows */
        }

        table tr:hover td {
            background-color: #2196F3 !important; /* Set background color of cells when hovered */
        }

    </style>
                     <table border="1" class="dataframe" style="width:100%;">
                                 <thead>
                                     <tr style="text-align: right;">\n"""
        k = 0
        counter_max = 0
        for i in ws.iter_rows():
            counter_max += 1

        for i in ws.iter_rows():
            if str(i[0].fill.start_color.rgb) != 'FFFFFFFF' or k == 0 or k == counter_max:
                if k > 0:
                    html += '<tr>\n'
                for j in range(len(i)):
                    value = str(i[j].value)
                    color = str(i[j].fill.start_color.rgb)
                    if (value == 'None'):
                        value = ' '
                    if color == '00000000':
                        color = 'FFFFFFFF'

                    if k < 1:
                        html += "<th style='background-color: #" + color[2:8] + "'>" + value + "</th>\n"
                    else:
                        html += "<td style='background-color: #" + color[2:8] + "'>" + value + "</td>\n"
                html += '</tr>\n'
                if k < 1:
                    html += """</thead>
                          <tbody>"""
                k += 1
        html += """ </tbody>
                    </table>
                    {% endblock %}"""
        html.encode('UTF-8')

        with open(name_destination + ".html", 'w', encoding='utf-8') as f:
            f.write(html)
            return True
    except:
        print(name_html)
        pass


create_folders(folder, folder_excels, folder_htmls)
rename_files(folder_excels)
for dirpath, dirnames, filenames in os.walk(folder_excels):
    for filename in filenames:
        if filename.endswith('.xlsx') and ('V00' in filename or 'K00' in filename):
            path = os.path.join(folder_excels, filename)
            name_dist = os.path.join(folder_htmls, filename[:-5])
            create_tables(path[:-5], name_dist)
