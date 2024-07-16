headers = [
    "RK",
    "GF-Nr",
    "Datum",
    "Vertreter",
    "ProvProz",
    "Provision",
    "Summe",
    "Lieferant",
    "Lieferantenname",
    "Name2",
    "IntKz",
    "PLZ Neu",
    "Ort",
    "s",
    "Paid amount",
    "Payment status",
    "Payment commission",
    "Payments with Rest",
    "Payments Dates"
]

CURRENT_DATE = "05.2024"
italy_agents = ["V000161","V000158","V000155",'V000147','V000142','V000122','V000157','V000030',"V000096",'V000136','V000184', 'V000186']
quartals = {"First":["01.2024","02.2024","03.2024"],"Second":["04.2024","05.2024","06.2024"],"Third":["07.2024","08.2024","09.2024"],"Fourth":["10.2024","11.2024","12.2024"]}
dict = {}

import openpyxl
from openpyxl.styles import PatternFill


def copy_cell_one_down(sheet, row, col):
    """
    Copy the value of a cell to the cell one row below it.

    :param sheet: The Excel worksheet instance.
    :param row: Row index of the cell to be copied.
    :param col: Column index of the cell to be copied.
    """
    source_cell = sheet.cell(row=row, column=col)
    target_cell = sheet.cell(row=row + 2, column=col)
    target_cell.value = source_cell.value
    target_cell.fill = openpyxl.styles.PatternFill(fill_type=source_cell.fill.fill_type,
                                                   fgColor=source_cell.fill.fgColor)
def color_cells(filename):
    # Load the Excel workbook and the active sheet
        wb = openpyxl.load_workbook(filename)

        sheet = wb.active
        t1=0
        # Define fills
        green_fill = PatternFill(start_color="00B300", end_color="00B300", fill_type="solid")
        light_green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        total_sum = 0.0
        comission_sum = 0.0
        # Iterate through the range O:Z for all rows in the sheet


        for col in range(19, 27):  # Columns O to Z have indexes 15 to 27
            for row in range(1, sheet.max_row + 1):  # Iterate over all rows
                cell = sheet.cell(row=row, column=col)
                cell_pecrent = sheet.cell(row=row, column=5)
                cell_requested = sheet.cell(row=row, column=7)
                if CURRENT_DATE in str(cell.value):
                    if cell.value and (",  " in str(cell.value) or ", " in str(cell.value)):
                        try:
                            float_val = float(str(cell.value).split(",  ")[1])
                            total_sum += float_val

                            percent = float(str(cell_pecrent.value))
                            comission_sum += float_val * percent/100
                        except:
                            pass

                    cell.fill = green_fill
                elif "2024" in str(cell.value):
                    cell.fill = light_green_fill

        # for row in range(1, sheet.max_row + 1):  # Iterate over all rows
        #         cell = sheet.cell(row=row, column=18)
        #         cell_pecrent = sheet.cell(row=row, column=5)
        #         if CURRENT_DATE in str(cell.value):
        #             if cell.value and (",  " in str(cell.value) or ", " in str(cell.value)):
        #                 try:
        #                     float_val = float(str(cell.value).split(",  ")[1])
        #                     total_sum -= float_val
        #                     percent = float(str(cell_pecrent.value))
        #                     comission_sum -= float_val * percent/100
        #                 except:
        #                     pass
                    # cell.fill = green_fill
                # elif "2024" in str(cell.value):
                    # cell.fill = light_green_fill

        for row in range(1, sheet.max_row + 1):  # Iterate over all rows
            cell = sheet.cell(row=row, column=15)
            try:
                float_val = float(str(cell.value))
                t1+= float_val
            except:
                pass
        comission_sum= round(comission_sum,2)
        copy_cell_one_down(sheet, sheet.max_row, 15)
        copy_cell_one_down(sheet, sheet.max_row-2, 17)

        source_cell2 = sheet.cell(row=sheet.max_row-2, column=14)
        target_cell2 = sheet.cell(row=sheet.max_row, column=14)
        source_cell3 = sheet.cell(row=sheet.max_row - 2, column=16)
        target_cell3 = sheet.cell(row=sheet.max_row, column=16)
        source_cell2.value = "September paid amount"
        target_cell2.value = "All paid amount"
        source_cell3.value = "September comissions"
        target_cell3.value = "All comissions"
        #target_cell.value = source_cell.value

        sheet.cell(row=sheet.max_row - 2, column=17, value=comission_sum)
        sheet.cell(row=sheet.max_row-2, column=15, value=total_sum)
        # sheet.cell(row=sheet.max_row + 2, column=1, value=t1)
        # Save the modified workbook
        print(comission_sum)
        wb.save(filename)
def make_agents_tables(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    # Extract all rows
    rows = list(sheet.iter_rows(values_only=True))

    # Skip header (if necessary)
    data = rows[1:]
    vertreter_dict = {}

    # Iterate over each row in data
    for row in data:
        vertreter_value = row[3]
        if vertreter_value not in vertreter_dict:
            vertreter_dict[vertreter_value] = []
        vertreter_dict[vertreter_value].append(row)
    fill1 = PatternFill(start_color="899ef0", end_color="899ef0", fill_type="solid")
    fill2 = PatternFill(start_color="00ff00", end_color="00ff00", fill_type="solid")
    fill3 = PatternFill(start_color="fff300", end_color="fff300", fill_type="solid")
    fill4 = PatternFill(start_color="899ef0", end_color="899ef0", fill_type="solid")
    fill5 = PatternFill(start_color="899ef0", end_color="899ef0", fill_type="solid")
    # Create sheets for each Vertreter
    for vertreter, vertreter_data in vertreter_dict.items():
        # Create a new sheet with the name as Vertreter value
        new_sheet = wb.create_sheet(title=str(vertreter) + " Total 24")

        # Write headers to the new sheet
        new_sheet.append(headers)

        # Set color for the header row
        for cell in new_sheet[1]:
            cell.fill = fill1

        # Write rows to the new sheet
        for row in vertreter_data:
            new_sheet.append(row)
            if str(new_sheet.cell(row=new_sheet.max_row, column=16).value) == "PAID" or str(new_sheet.cell(row=new_sheet.max_row, column=16).value) == "PARTIALLY PAID":
                last_column = new_sheet.max_column
                for col_num in range(1,17):
                    new_sheet.cell(row=new_sheet.max_row, column=col_num).fill = fill2
            if str(new_sheet.cell(row=new_sheet.max_row, column=16).value) == "CREDIT NOTE":
                last_column = new_sheet.max_column
                for col_num in range(1,17):
                    new_sheet.cell(row=new_sheet.max_row, column=col_num).fill = fill3


#ПОДСЧЕТ ТУТ
    try:
        for i in range(2, len(wb.worksheets)):
            sheet = wb.worksheets[i]
            total_sum = 0
            comission_sum = 0
            for row in range(2, sheet.max_row + 1):  # Iterate over all rows
                row_sum = 0
                percent_sum = 0

                cell = sheet.cell(row=row, column=7)
                print(cell.value)
                all_value = float(str(cell.value))
                all_percent_value = 0

                for col in range(19, 27):  # Columns O to Z have indexes 15 to 27

                    cell = sheet.cell(row=row, column=col)
                    cell_pecrent = sheet.cell(row=row, column=5)
                    if cell.value and (",  " in str(cell.value) or ", " in str(cell.value)):
                        try:
                            float_val = float(str(cell.value).split(",  ")[1])
                            row_sum += float_val
                            percent = float(str(cell_pecrent.value))
                            percent_sum += float_val * percent / 100
                        except:
                            pass

                    if col == 26:
                        if all_value > 0:
                            if all_value>row_sum:
                                total_sum += row_sum
                                comission_sum += percent_sum
                            else:
                                try:
                                    total_sum+=all_value
                                    comission_sum += all_value * float(str(cell_pecrent.value)) / 100
                                except:pass
            # for row in range(1, sheet.max_row + 1):  # Iterate over all rows
            #     cell = sheet.cell(row=row, column=18)
            #     cell_pecrent = sheet.cell(row=row, column=5)
            #     if cell.value and (",  " in str(cell.value) or ", " in str(cell.value)):
            #         try:
            #             float_val = float(str(cell.value).split(",  ")[1])
            #             total_sum -= float_val
            #             percent = float(str(cell_pecrent.value))
            #             comission_sum -= float_val * percent / 100
            #         except:
            #             pass
            source_cell2 = sheet.cell(row=sheet.max_row + 1, column=15)
            source_cell3 = sheet.cell(row=sheet.max_row, column=14)
            total_sum = round(total_sum, 2)
            source_cell2.value = total_sum
            source_cell3.value = "All paid amount"

            source_cell4 = sheet.cell(row=sheet.max_row, column=17)
            source_cell5 = sheet.cell(row=sheet.max_row, column=16)
            comission_sum = round(comission_sum, 2)
            source_cell4.value = comission_sum
            source_cell5.value = "All comissions"
            light_green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

            source_cell2.fill = light_green_fill
            source_cell4.fill = light_green_fill
    except:
        pass
    # Save the workbook with the new sheets
    wb.save('updated_filename_fix.xlsx')
    # Print the result
    print(vertreter_dict)
    pass
def make_month_tables(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    # Extract all rows
    rows = list(sheet.iter_rows(values_only=True))

    # Skip header (if necessary)
    data = rows[1:]
    vertreter_dict = {}

    # Iterate over each row in data
    for row in data:
        vertreter_value = row[3]
        row1 = []
        paids = []
        paid = 0
        comission = 0
        at_least_one = True

        for x in range(18, len(row)):
            if vertreter_value in italy_agents:
                for date in quartals['Second']:
                    if date in str(row[x]):
                        if row[x] and (",  " in str(row[x]) or ", " in str(row[x])):
                            try:
                                float_val = float(str(row[x]).split(",  ")[1])
                                paid += float_val
                                paids.append(row[x])
                                percent = float(str(row[4]))
                                comission += float_val * percent/100
                            except:
                                pass
            else:
                if CURRENT_DATE in str(row[x]):
                    if row[x] and (",  " in str(row[x]) or ", " in str(row[x])):
                        try:
                            float_val = float(str(row[x]).split(",  ")[1])
                            paid += float_val
                            paids.append(row[x])
                            percent = float(str(row[4]))
                            comission += float_val * percent/100
                        except:
                            pass




        if len(paids) == 0:
            at_least_one = False
        else:
            for x in range(18):
                row1.append(row[x])
            for x in range(len(paids)):
                row1.append(paids[x])
            row1[2] = str(row1[2]).replace(" 00:00:00","")
            row1[14] = paid
            print(float(row1[6]))
            print(paid)
            # print()
            if paid == 0:
                row1[15] = "UNPAID"
            elif paid / float(row1[6]) >= 1:
                row1[15] = "PAID"
            else:
                row1[15] = "PARTIALLY PAID"
            row1[16] = round(comission,1)
            if vertreter_value in italy_agents:
                for date in quartals['Second']:
                    if date in str(row1[17]):
                        break
                    else:
                        row1[17] = None
            else:
                if CURRENT_DATE in str(row1[17]):
                    pass
                else:
                    row1[17] = None




        if row[15] == "CREDIT NOTE" :
            for x in range(18):
                row1.append(row[x])
            row1[2] = str(row1[2]).replace(" 00:00:00","")
            at_least_one = True
        if at_least_one:
            if vertreter_value not in vertreter_dict:
                vertreter_dict[vertreter_value] = []
            vertreter_dict[vertreter_value].append(tuple(row1))

    fill1 = PatternFill(start_color="899ef0", end_color="899ef0", fill_type="solid")
    fill2 = PatternFill(start_color="00ff00", end_color="00ff00", fill_type="solid")
    fill3 = PatternFill(start_color="fff300", end_color="fff300", fill_type="solid")



    # Create sheets for each Vertreter
    for vertreter, vertreter_data in vertreter_dict.items():
        # Create a new sheet with the name as Vertreter value
        new_sheet = ""
        if vertreter in italy_agents:
            new_sheet = wb.create_sheet(title=str(vertreter) + " Second 24")
        else:
            new_sheet = wb.create_sheet(title=str(vertreter) + " May 24")

        # Write headers to the new sheet
        new_sheet.append(headers)

        # Set color for the header row
        for cell in new_sheet[1]:
            cell.fill = fill1

        # Write rows to the new sheet
        for row in vertreter_data:
            new_sheet.append(row)
            if str(new_sheet.cell(row=new_sheet.max_row, column=16).value) == "PAID" or str(new_sheet.cell(row=new_sheet.max_row, column=16).value) == "PARTIALLY PAID":
                last_column = new_sheet.max_column
                for col_num in range(1,17):
                    new_sheet.cell(row=new_sheet.max_row, column=col_num).fill = fill2
            if str(new_sheet.cell(row=new_sheet.max_row, column=16).value) == "CREDIT NOTE":
                last_column = new_sheet.max_column
                for col_num in range(1,17):
                    new_sheet.cell(row=new_sheet.max_row, column=col_num).fill = fill3

    for i in range(2, len(wb.worksheets)):
        sheet = wb.worksheets[i]
        total_sum = 0
        comission_sum = 0
        for col in range(19, 27):  # Columns O to Z have indexes 15 to 27
            for row in range(1, sheet.max_row + 1):  # Iterate over all rows
                cell = sheet.cell(row=row, column=col)
                cell_pecrent = sheet.cell(row=row, column=5)
                if cell.value and (",  " in str(cell.value) or ", " in str(cell.value)):
                    try:
                        float_val = float(str(cell.value).split(",  ")[1])
                        total_sum += float_val
                        percent = float(str(cell_pecrent.value))
                        comission_sum += float_val * percent / 100
                    except:
                        pass
        # for row in range(1, sheet.max_row + 1):  # Iterate over all rows
        #     cell = sheet.cell(row=row, column=18)
        #     cell_pecrent = sheet.cell(row=row, column=5)
        #     if cell.value and (",  " in str(cell.value) or ", " in str(cell.value)):
        #         try:
        #                     float_val = float(str(cell.value).split(",  ")[1])
        #                     total_sum -= float_val
        #                     percent = float(str(cell_pecrent.value))
        #                     comission_sum -= float_val * percent/100
        #         except:
        #                     pass
        source_cell2 = sheet.cell(row=sheet.max_row+1, column=15)
        source_cell3 = sheet.cell(row=sheet.max_row, column=14)
        total_sum = round(total_sum,2)
        source_cell2.value = total_sum
        source_cell3.value = "All paid amount"

        source_cell4 = sheet.cell(row=sheet.max_row , column=17)
        source_cell5 = sheet.cell(row=sheet.max_row, column=16)
        comission_sum= round(comission_sum,2)
        source_cell4.value = comission_sum
        source_cell5.value = "All comissions"
        light_green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

        source_cell2.fill = light_green_fill
        source_cell4.fill = light_green_fill
    # Save the workbook with the new sheets
    wb.save('updated_filename2.xlsx')
    # Print the result
    print(vertreter_dict)
    pass


if __name__ == "__main__":
    filename = "may_2024/OutputResult0.xlsx"
    # color_cells(filename)
    # make_agents_tables(filename)
    make_month_tables(filename)
    # color_cells('october/updated_filename2_old.xlsx')
