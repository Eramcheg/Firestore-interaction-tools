import firebase_admin
from firebase_admin import credentials, firestore
from openpyxl import load_workbook

# Инициализация Firebase Admin SDK
cred = credentials.Certificate("G:\\FIles\\firebase\\key2.json")
firebase_admin.initialize_app(cred)

# Подключение к Firestore
db = firestore.client()
collection_ref = db.collection("Stones")

# Путь к Excel-файлу
file_path = '..\\static_files\\stones.xlsx'

# Чтение Excel-файла с использованием openpyxl
wb = load_workbook(file_path)
sheet = wb.active

# Проверяем, что первая строка содержит заголовки: "id" и "name"
header = [cell.value for cell in sheet[1]]
if "id" not in header or "name" not in header:
    raise ValueError("Excel file must contain 'id' and 'name' columns.")

# Индексы колонок "id" и "name"
id_index = header.index("id") + 1
name_index = header.index("name") + 1

# Итерация по строкам таблицы, начиная со второй строки
for row in sheet.iter_rows(min_row=2, values_only=True):
    stone_id = row[id_index - 1]
    stone_name = row[name_index - 1]

    if not stone_id or not stone_name:
        print(f"Skipping row with missing data: {row}")
        continue

    # Добавление данных в Firestore
    collection_ref.document(str(stone_id)).set({
        'id': str(stone_id),
        'name': str(stone_name)
    })
    print(f"Added: {stone_id} - {stone_name}")

print("Data import completed!")