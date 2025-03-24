from openpyxl import Workbook
import re
from openpyxl import load_workbook

# Путь к существующему Excel файлу
file_path = '../static_files/all_items06012025.xlsx'

# Открываем Excel файл
workbook = load_workbook(file_path)
worksheet = workbook.active

# Читаем заголовки (первую строку таблицы)
headers = [cell.value for cell in worksheet[1]]

# Определяем индексы нужных колонок
name_idx = headers.index("name")
visible_idx = headers.index("Visible")
stone_idx = headers.index("stone")
category_idx = headers.index("category")
product_name_idx = headers.index("product_name")
description_idx = headers.index("description")
material_idx = headers.index("material")
image_url_idx = headers.index("image_url")
collection_idx = headers.index("collection") if "collection" in headers else None
pre_order_idx = headers.index("pre_order") if "pre_order" in headers else None

# Регулярное выражение для проверки имени товара
regex = re.compile(r"^[a-zA-Z]{0,2}\d{5}[a-zA-Z]{0,5}$")

# Пример словаря stones для проверки
stones = {"diamond": "Diamond", "ruby": "Ruby"}  # Пример, замените на реальные данные

# Функция для расчета цены (заглушка)
def calculate_price(item, price_category, sale):
    # Замените логику на свою
    return 100.0  # Примерное значение

# Итоговый словарь для хранения товаров
products = {}

# Проход по строкам в таблице (начиная со второй)
for row in worksheet.iter_rows(min_row=2, values_only=True):
    if not row[visible_idx]:  # Проверяем, что товар видимый
        continue

    item_name = row[name_idx]
    stone = stones.get(row[stone_idx], row[stone_idx])
    item_name = item_name.split(' ')[0]  # Берём первую часть имени
    regex1 = re.compile(r"^[a-zA-Z]{0,2}\d{5}")
    match = regex1.match(item_name)  # Проверяем соответствие регулярному выражению
    if match:
        # Извлекаем первые буквы и цифры (объединяем первые две группы)
        item_name = match.group(0)
    else:
        # Если не соответствует регулярке, оставляем первую часть имени
        item_name = row[name_idx].split(' ')[0]

    # Добавляем товар в словарь, если его там еще нет
    if item_name not in products:
        products[item_name] = {
            "name": item_name,
            "product_name": f"{row[product_name_idx]}",
            "description": row[description_idx],
            "price": calculate_price(row, None, None),  # Укажите цену, если есть категории или скидки
            "category": row[category_idx],
            "material": row[material_idx],
            "preview_image": row[image_url_idx],
        }
    else:
        print(f"Item {item_name} already exists in the dictionary")

# Создание нового Excel файла
new_workbook = Workbook()
new_worksheet = new_workbook.active
new_worksheet.title = "Processed Items"

# Запись заголовков в новый файл
new_headers = ["name", "product_name", "description", "price", "category", "material", "preview_image"]
new_worksheet.append(new_headers)

# Запись данных из словаря products в новый файл
for product_data in products.values():
    new_worksheet.append([
        product_data["name"],
        product_data["product_name"],
        product_data["description"],
        product_data["price"],
        product_data["category"],
        product_data["material"],
        product_data["preview_image"],
    ])

# Сохранение нового файла
new_file_path = '../static_files/processed_items.xlsx'
new_workbook.save(new_file_path)
print(f"Данные успешно сохранены в {new_file_path}")
