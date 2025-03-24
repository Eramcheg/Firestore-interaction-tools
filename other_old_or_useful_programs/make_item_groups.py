from openpyxl import load_workbook

# Путь к существующему Excel файлу
file_path = '../static_files/processed_items.xlsx'

# Открываем Excel файл
workbook = load_workbook(file_path)
worksheet = workbook.active

# Читаем заголовки (первую строку таблицы)
headers = [cell.value for cell in worksheet[1]]

# Определяем индексы интересующих полей
product_name_idx = headers.index("product_name")  # Замените "product_name" на точное название столбца
name_idx = headers.index("name")  # Замените "name" на точное название столбца

# Создаем словарь для группировки
grouped_products = {}

# Проходим по строкам начиная со второй (первая — заголовки)
for row in worksheet.iter_rows(min_row=2, values_only=True):
    product_name = row[product_name_idx]
    name = row[name_idx]

    if product_name and name:  # Проверяем, что оба значения существуют
        if product_name not in grouped_products:
            grouped_products[product_name] = []
        grouped_products[product_name].append(name)

# Результат: grouped_products содержит группы товаров
print(grouped_products)