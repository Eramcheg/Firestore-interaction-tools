import openpyxl
import csv
import io
import firebase_admin
from firebase_admin import credentials, firestore
cred = credentials.Certificate("G:\\FIles\\firebase\\key2.json")
firebase_admin.initialize_app(cred)
db = firestore.client()
collection_ref = firestore.client().collection("item")

workbook = openpyxl.load_workbook("../../../agents app/storages/01.04.2025.xlsx")
sheet = workbook.active
# Function to convert to float or return 0
def to_float(value):
    try:
        return float(value)
    except ValueError:
        print(f"Invalid value: {value}")
        return 0
    except TypeError:
        print(f"Invalid value: {value}")
        return 0


# Step 1: Build the mapping from 'name' to document ID
name_to_doc_id = {}
docs = collection_ref.stream()
for doc in docs:
    doc_data = doc.to_dict()
    if 'name' in doc_data:
        name_to_doc_id[doc_data['name']] = doc.id

# Initialize variables for batching
batch = db.batch()
counter = 0
batch_counter = 0

# Step 2: Process the Excel sheet and update documents using the mapping
for row in range(2, sheet.max_row + 1):  # Пропускаем заголовки (первая строка)
    doc_name = sheet.cell(row=row, column=5).value  # Получаем имя документа из колонки E
    if doc_name in name_to_doc_id:
        doc_id = name_to_doc_id[doc_name]
        doc_ref = collection_ref.document(doc_id)
        print(f"Проверка документа {doc_name}")

        # Получаем текущие данные документа из Firestore
        current_data = doc_ref.get().to_dict()

        # Готовим словарь обновления только с изменёнными полями
        update_fields = {}

        new_price = to_float(sheet.cell(row=row, column=24).value)  # Колонка X
        if current_data.get("price") != new_price:
            update_fields["price"] = new_price

        new_priceVK4 = to_float(sheet.cell(row=row, column=25).value)  # Колонка Y
        if current_data.get("priceVK4") != new_priceVK4:
            update_fields["priceVK4"] = new_priceVK4

        # Здесь, как в примере, используется то же значение для priceVK3, что и для price
        new_priceVK3 = to_float(sheet.cell(row=row, column=24).value)  # Колонка X
        if current_data.get("priceVK3") != new_priceVK3:
            update_fields["priceVK3"] = new_priceVK3

        new_priceUSD_GH = to_float(sheet.cell(row=row, column=26).value)  # Колонка Z
        if current_data.get("priceUSD_GH") != new_priceUSD_GH:
            update_fields["priceUSD_GH"] = new_priceUSD_GH

        # Обратите внимание: для priceGH используется колонка 16 (проверьте, что это соответствует вашим данным)
        new_priceGH = to_float(sheet.cell(row=row, column=16).value)
        if current_data.get("priceGH") != new_priceGH:
            update_fields["priceGH"] = new_priceGH

        new_priceUSD = to_float(sheet.cell(row=row, column=28).value)  # Колонка AB
        if current_data.get("priceUSD") != new_priceUSD:
            update_fields["priceUSD"] = new_priceUSD

        # Если есть изменения, добавляем обновление в batch
        if update_fields:
            print(f"Обновляем документ {doc_name} с изменениями: {update_fields}")
            batch.update(doc_ref, update_fields)
            counter += 1
            batch_counter += 1

        # Commit batch every 500 updates or the final batch
        if batch_counter == 500:
            batch.commit()
            batch = db.batch()  # Start a new batch
            batch_counter = 0
            print(f"Processed and committed {counter} updates so far...")

# Commit any remaining updates in the batch
if batch_counter > 0:
    batch.commit()

print(f"Update complete. Total documents updated: {counter}.")