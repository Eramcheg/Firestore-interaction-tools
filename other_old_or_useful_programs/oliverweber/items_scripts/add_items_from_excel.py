import openpyxl
from openpyxl.workbook import Workbook
import datetime
import firebase_admin
from firebase_admin import credentials, firestore

# Инициализация Firebase (если необходимо)
cred = credentials.Certificate("G:\\FIles\\firebase\\key2.json")
firebase_admin.initialize_app(cred)
db = firestore.client()
collection_ref = db.collection("item")

# Путь к исходному файлу Excel
xlsx_file_path = "../../../static_files/18.03.2025.xlsx"
workbook = openpyxl.load_workbook(xlsx_file_path)
sheet = workbook.active

# Массив номеров для поиска
numbers = [
    "61322", "61324G", "62290", "62292G", "61320G", "61323G", "61321", "61321G", "61320", "62291G", "62289", "62289G",
    "12491G", "12493G", "12492G", "23214G", "23215G", "23216G", "12490G", "23213G", "41229S", "41229M","41229L", "41229XL", "41230S", "41230L", "41230M", "41230XL",
    "32432G", "12482G", "23205G", "32431", "23204", "12484", "23207", "23048", "12481",
    "12483G", "32433G", "23208G", "23206G", "12485G",
    "12489", "12488", "12487", "23212", "23211", "23210", "12486G", "12486", "23209G", "23209",
    "12157 L169I", "12157 L167I", "12157 L168I", "12157 L163I", "22916 L169I", "22916 L167I", "22916 L168I", "22916 L163I", "11522 971S", "11522 214", "22388 971S", "22388 214",
    "11523G 971S", "11523 1405", "11523 1201", "11523G 1002", "22389G 971S", "22389 1405", "22389 1201", "22389G 1002", "12052R 263", "12052R 238", "12052R 277", "22851R 263", "22851R 238", "22851R 277",
    "21009 L110D", "21009 1405", "21009 L134I", "21009 L107D", "21009 L169I", "21009 L142D", "21009 L167I", "21009 L163I", "21009 L168I", "21009 1201", "21009 L137I", "21009 L161I", "63322L BLU", "63322S BLU", "63322M BLU", "63322XL BLU"
]

# Заголовки для нового файла (колонки с характеристиками)
headers = [
    "Visible", "category", "description", "image-url", "image_url", "material", "name", "plating",
    "price", "priceGH", "priceUSD", "priceUSD_GH", "priceVK3", "priceVK4", "ean_13", "size",
    "product_name", "product_width", "product_height", "chain_length", "quantity", "stone"
]

# Значение по умолчанию для изображения
default_image = "https://firebasestorage.googleapis.com/v0/b/flutterapp-fd5c3.appspot.com/o/wall%2Fno_image.jpg?alt=media&token=22a7b907-01f6-45b6-8fb1-f1f884ab21d4"

# Список, куда будут сохраняться найденные продукты
products = []

# Для каждого номера ищем первую строку в исходном Excel, где в столбце с номером (индекс 4) значение совпадает
for num in numbers:
    product_found = False
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Проверяем, совпадает ли значение в ячейке с индексом 4 с номером (предполагается, что номер продукта хранится именно там)
        if row[4] == num:
            # Извлекаем необходимые данные. Индексы подобраны согласно вашему примеру.
            # Если некоторые поля отсутствуют в файле, можно оставить их пустыми.
            try:
                price = round(row[24], 2) if row[23] is not None else 0.0
                priceGH = round(row[16], 2) if row[15] is not None else 0.0
                priceUSD = round(row[27], 2) if row[26] is not None else 0.0
                priceUSD_GH = round(row[26], 2) if row[25] is not None else 0.0
                priceVK3 = round(row[24], 2) if row[23] is not None else 0.0
                priceVK4 = round(row[25], 2) if row[24] is not None else 0.0
            except Exception as e:
                price = priceGH = priceUSD = priceUSD_GH = priceVK3 = priceVK4 = 0.0

            product = {
                "Visible": "True",
                "category": "",
                "description": row[6],
                "image-url": default_image,
                "image_url": default_image,
                "material": "" ,
                "name": row[4],
                "plating": "",  # если информация о покрытии есть в файле, укажите нужный индекс
                "price": price,
                "priceGH": priceGH,
                "priceUSD": priceUSD,
                "priceUSD_GH": priceUSD_GH,
                "priceVK3": priceVK3,
                "priceVK4": priceVK4,
                "ean_13": row[7] if len(row) > 7 and row[7] is not None else "",
                "size":"",  # замените ? на индекс для размера, если он есть
                "product_name": "",  # аналогично для названия продукта
                "product_width": row[19],
                "product_height": row[18],
                "chain_length": row[21],
                "quantity": row[2],
                "stone": ""  # если есть информация о камнях, укажите индекс
            }
            products.append(product)
            product_found = True
            break
    # Если номер не найден в исходном файле, можно добавить запись с номером и пустыми характеристиками
    if not product_found:
        product = {
            "Visible": True,
            "category": "WHOLE PRODUCT NOT FOUND",
            "description": "",
            "image-url": default_image,
            "image_url": default_image,
            "material": "",
            "name": num,
            "plating": "",
            "price": 0.0,
            "priceGH": 0.0,
            "priceUSD": 0.0,
            "priceUSD_GH": 0.0,
            "priceVK3": 0.0,
            "priceVK4": 0.0,
            "ean_13": "",
            "size": "",
            "product_name": "",
            "product_width": "",
            "product_height": "",
            "chain_length": "",
            "quantity": 0,
            "stone": ""
        }
        products.append(product)

# Создаём новый Excel-файл с найденными данными
wb_new = Workbook()
ws_new = wb_new.active
ws_new.title = "Products"

# Записываем заголовки
ws_new.append(headers)

# Записываем данные по каждому продукту
for prod in products:
    row_data = [prod[h] for h in headers]
    ws_new.append(row_data)

output_file = "../../../static_files/springproducts.xlsx"
wb_new.save(output_file)
print(f"Файл сохранён: {output_file}")
