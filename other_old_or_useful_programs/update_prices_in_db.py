import openpyxl
import csv
import io
import firebase_admin
from firebase_admin import credentials, firestore
cred = credentials.Certificate("G:\\FIles\\firebase\\key2.json")
firebase_admin.initialize_app(cred)
db = firestore.client()
collection_ref = firestore.client().collection("item")

workbook = openpyxl.load_workbook("../agents app/storages/05.04.2024.xlsx")
sheet = workbook.active
# Function to convert to float or return 0
def to_float(value):
    try:
        return float(value)
    except ValueError:
        print(f"Invalid value: {value}")
        return 0
    except TypeError:
        print(f"Invalid value: {value}")
        return 0


# Step 1: Build the mapping from 'name' to document ID
name_to_doc_id = {}
docs = collection_ref.stream()
for doc in docs:
    doc_data = doc.to_dict()
    if 'name' in doc_data:
        name_to_doc_id[doc_data['name']] = doc.id

# Initialize variables for batching
batch = db.batch()
counter = 0
batch_counter = 0

# Step 2: Process the Excel sheet and update documents using the mapping
for row in range(2, sheet.max_row + 1):  # Assuming the first row is headers
    doc_name = sheet.cell(row=row, column=5).value  # Column E
    if doc_name in name_to_doc_id:
        doc_id = name_to_doc_id[doc_name]
        doc_ref = collection_ref.document(doc_id)
        print(f"Updating document {doc_name}")
        # Prepare batch update
        batch.update(doc_ref, {
            "price": to_float(sheet.cell(row=row, column=24).value),  # Column X
            "priceVK4": to_float(sheet.cell(row=row, column=25).value),  # Column Y
            "priceVK3": to_float(sheet.cell(row=row, column=24).value),  # Column X
        })
        counter += 1
        batch_counter += 1

        # Commit batch every 500 updates or the final batch
        if batch_counter == 500:
            batch.commit()
            batch = db.batch()  # Start a new batch
            batch_counter = 0
            print(f"Processed and committed {counter} updates so far...")

# Commit any remaining updates in the batch
if batch_counter > 0:
    batch.commit()

print(f"Update complete. Total documents updated: {counter}.")