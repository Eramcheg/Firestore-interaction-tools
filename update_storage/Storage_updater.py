# import csv
# import firebase_admin
# from firebase_admin import credentials, firestore
#
# # Initialize Firebase Admin with your service account credentials
# cred = credentials.Certificate("G:\\FIles\\firebase\\key2.json")
# firebase_admin.initialize_app(cred)
#
# # Reference the collection where you want to update the document
# collection_ref = firestore.client().collection("item")
#
# # Specify the field and its value for the query
# path = "storage.csv"
#
# # Prepare a batch for updates
# batch = firestore.client().batch()
#
# with open(path, 'r') as file:
#     csv_reader = csv.reader(file, delimiter=';')
#     for row in csv_reader:
#         if row[1] != 'quantity':
#             query_value = row[0]
#             try:
#                 new_value = int(row[1])
#             except:
#
#                 value = round(float(row[1]) + ((-0.5) if float(row[1])>0 else 0.5), 0 )
#                 # quantity.append(int(value))
#                 new_value = value
#             #new_value = row[1]
#
#             # Query for documents where the specified field matches the specified value
#             query = collection_ref.where("name", "==", query_value)
#             query_results = query.get()
#
#             for doc_snapshot in query_results:
#                 # Update the document in the batch
#
#                 field1_value = doc_snapshot.get("quantity")
#                 print("Olda value:", field1_value, "New value:", new_value)
#                 if field1_value != new_value:
#                     doc_ref = doc_snapshot.reference
#                     batch.update(doc_ref, {"quantity": new_value})
#
#
# # Commit the batch updates
# batch.commit()
import openpyxl
import csv
import io
import firebase_admin
from firebase_admin import credentials, firestore

columnQuantity = 2  # Starts with 0
columnNumber = 4 # Starts with 0


# Initialize Firebase Admin with your service account credentials
cred = credentials.Certificate("G:\\FIles\\firebase\\key2.json")
firebase_admin.initialize_app(cred)

# Reference the collection where you want to update the document
collection_ref = firestore.client().collection("item")

# Load Excel workbook
xlsx_file_path = "../storages/04.01.2023.xlsx"
workbook = openpyxl.load_workbook(xlsx_file_path)
sheet = workbook.active

col = 0
for i in sheet.iter_cols():
    if(i[0].value == 'Lagerstand Gesamt'):
        columnQuantity=col
    if(i[0].value == 'Artikelnummer'):
        columnNumber=col
    print(i[0].value)
    col+=1


# Create an in-memory text stream for CSV data
output = io.StringIO()
csv_writer = csv.writer(output, delimiter=';')

# Write header to CSV data
csv_writer.writerow(['name', 'quantity'])

# Iterate through rows in the XLSX sheet and extract data
for row in sheet.iter_rows(min_row=2, values_only=True):
    name = row[columnNumber]  # Assuming 'name' is in the 5th column
    quantity = row[columnQuantity] if row[2] is not None else 0  # Assuming 'quantity' is in the 3rd column
    quantity = max(int(round(quantity, 0)), 0)  # Ensure non-negative quantity
    csv_writer.writerow([name, quantity])

# Set the pointer of the in-memory text stream to the beginning
output.seek(0)

# Prepare a batch for updates
batch = firestore.client().batch()
# print(output)
# Process the in-memory CSV data
csv_reader = csv.reader(output, delimiter=';')
next(csv_reader)  # Skip the header row
for row in csv_reader:
    query_value = row[0]
    try:
        new_value = int(row[1])
    except ValueError:
        # Handle float to int conversion with rounding
        value = round(float(row[1]) + (-0.5 if float(row[1]) > 0 else 0.5), 0)
        new_value = int(value)

    # Query for documents
    query = collection_ref.where("name", "==", query_value)
    query_results = query.get()

    for doc_snapshot in query_results:
        # Update the document in the batch
        field1_value = doc_snapshot.get("quantity")
        if field1_value != new_value:
            print(f"Old value:{field1_value}    New value: {new_value}")
            doc_ref = doc_snapshot.reference
            batch.update(doc_ref, {"quantity": new_value})

# Commit the batch updates

batch.commit()
print('All values were updated')