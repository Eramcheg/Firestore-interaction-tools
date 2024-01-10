import datetime

import openpyxl

italian_names = {
    "Anhänger":"Pendente",
    "Armband":"Braccialetto",
    "Ohrringe": "Orrechini",
    "Ohrstecker": "Orrechini",
    "Ohrhänger": "Orrechini",
    "Ring":"Anello",
    "Armreif": "Braccialetto",
    "Fusskette": "Braccialetto",
    "Kette": "Pendente",

}

def search_and_create_new_excel(input_strings, source_excel_file, output_excel_file, group_num):
    # Load the source Excel file
    wb = openpyxl.load_workbook(source_excel_file)
    sheet = wb.active

    # Create a new Excel file
    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active
    new_sheet.title = group_num
    print('here')
    header =["FA_SORTNR","Artnr.","best. Menge","Artikelbezeichnung","VK4"]
    # Iterate over input strings and find matching rows
    new_sheet.append(header)
    for index, input_string in enumerate(input_strings, start=1):
        for row in sheet.iter_rows(min_row=2):  # Assuming the first row is a header
            if row[4].value == input_string:

                description = row[6].value
                for german_name, italian_name in italian_names.items():
                    if german_name in description:
                        description = description.replace(german_name, italian_name)
                        break


                new_sheet.append([
                    index,                    # Ordinal number
                    input_string,             # Input string
                    row[2].value,                       # Empty column //38
                    description,             # 5th column from original file
                    row[15].value # 47th column from original file  //46
                ])
                break
            else:
                continue


    # Save the new Excel file
    new_wb.save(f"new_vitrines/{output_excel_file}")
print("Type name of your file. Press enter if you'd like to use default name")
file_name = input()


print("Provide group number first, then provide product's numbers")
group_num = input()
while not group_num.isdigit():
    print("Provide valid group number")
    group_num = input()

today = datetime.date.today()
if file_name == "":
    file_name = f"v{group_num}_{today.strftime('%d_%m_%y')}.xlsx"

print(f"You entered {group_num}. \nNow provide numbers, each from new line, or type -1 to finish")

# Example usage
input_strings = []
inp = ""
while inp!="-1":
    inp = input()
    print(inp)
    if(inp != "-1"):
        print(f"inside {inp}")
        input_strings.append(inp)

print(input_strings)
search_and_create_new_excel(input_strings, 'articles.xlsx', file_name, f"v{group_num}")